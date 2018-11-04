# 두 엑셀파일 읽어 하나로 합치기
import pandas as pd

df_삼성전자 = pd.read_excel('2017년 광고비 - 삼성전자.xlsx')
df_삼성전자.set_index('date', inplace=True)

df_LG전자 = pd.read_excel('2017년 광고비 - LG전자.xlsx')
df_LG전자.set_index('date', inplace=True)

df_merge = pd.DataFrame()

df_merge['삼성전자'] = df_삼성전자['total']
df_merge['LG전자'] = df_LG전자['total']

df_merge.to_excel('merged_01.xlsx')

# 계산 및 계산결과 저장
import openpyxl  

    # 엑셀 파일 열고, 시트 얻기
wb = openpyxl.load_workbook('merged_01.xlsx')
sheet = wb.active
sheet

    # 삼성전자 컬럼
삼성전자_월광고비 = [row[0].value for row in sheet['B2':'B13']]
삼성전자_월광고비

    # 합계
sum(삼성전자_월광고비)

sheet['B14'].value = sum(삼성전자_월광고비)
sheet['B14'].value

    # '합계'
sheet['A14'].value = '합계'

    # 합계 계산
sheet['B14'].value = sum([row[0].value for row in sheet['B2':'B13']])
sheet['C14'].value = sum([row[0].value for row in sheet['C2':'C13']])

wb.save("merged_02.xlsx")

# 수식 넣기
    # 합계 계산
sheet['B14'].value = '=SUM(B2:B13)'
sheet['C14'].value = '=SUM(C2:C13)'

# 셀에 스타일 지정
from openpyxl.styles import Font, Alignment, Border, Side, Color, PatternFill

    # Font: '맑은 고딕', 크기 15, 굵게
font_15 = Font(name='맑은 고딕', size=15, bold=True)

    # Alignment: 가로 세로 , 가운데 정렬
align_center = Alignment(horizontal='center', vertical='center')
align_vcenter = Alignment(vertical='center')

    # Border: 테두리 지정
border_thin = Border(
    left=Side(style='thin'), right=Side(style='thin'), 
    top=Side(style='thin'), bottom=Side(style='thin')
)

    # PatternFill: 셀 색상 지정
fill_orange = PatternFill(patternType='solid', fgColor=Color('FFC000'))
fill_lightgrey = PatternFill(patternType='solid', fgColor=Color('D3D3D3'))

cell_sum = sheet['A14'] # 합계 제목 셀

cell_sum.font = font_15
cell_sum.alignment = align_center
cell_sum.border = border_thin
cell_sum.fill = fill_orange

wb.save("merged_04.xlsx")

# 범위에 스타일 지정
    # 범위(range)
sheet['B2:C14']

for row in sheet['B2:C14']:
    for cell in row:
        cell.border = border_thin
        cell.number_format = '0.00'
        
for row in sheet['B14:C14']:
    for cell in row:
        cell.alignment = align_vcenter
        cell.fill = fill_orange

wb.save("merged_05.xlsx")

# 차트 추가하기
from openpyxl.chart import Reference, Series, BarChart

chart = BarChart()
chart.title = '2017년 월별 광고비 (억원)'

values = Reference(sheet, range_string= 'Sheet1!B1:B13')
series = Series(values, title="삼성전자")
chart.append(series)

values = Reference(sheet, range_string= 'Sheet1!C1:C13')
series = Series(values, title="LG전자")
chart.append(series)

sheet.add_chart(chart, 'E1')

wb.save("merged_06.xlsx")

